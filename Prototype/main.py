# modules
import sys
import mysql.connector
import openpyxl


from openpyxl import Workbook, load_workbook
from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QWidget, QCheckBox, QGridLayout, QStyledItemDelegate, QHBoxLayout
from PyQt5.QtCore import pyqtSlot, QFile, QTextStream
from PyQt5 import QtCore, QtGui, QtWidgets

# import files
from sidebar_ui import Ui_MainWindow
from product_ui import Ui_Form
from do_form_ui import Ui_Form as DoFormUi
from invoice_form_ui import Ui_Form as InvoiceFormUi


# class
class ReadOnlyDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):
        print('editing prohibited!')
        return

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        # Load the UI from the sidebar_ui module
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.setWindowTitle("TDB App")

        ## setup title
        # header for table page
        self.ui.po_table.setHorizontalHeaderLabels(["View Products", "Purchase No", "Purchase Date", "Product Qty", "Make DO"])
        self.ui.products_table.setHorizontalHeaderLabels(["" , "", "Material No", "Nama Produk", "Qty", "UoM", "Harga Unit", "Mata Uang" ,"Discount (%)", "Total Harga"])
        self.ui.do_table.setHorizontalHeaderLabels(["Invoice NO", "", "More Detail" , "Delivery No", "Delivery Date", "Purchase No", "Product Qty", "Note", "Invoice Status", "Make Invoice"])
        self.ui.do_view_more_table.setHorizontalHeaderLabels(["Product_id", "", "Material No", "Nama Produk", "Product", "Qty", "UoM", "Remark"])

        # sql connection
        self.connection = mysql.connector.connect (
            host="localhost",
            user="root",
            password="",
            database="database_tdb"
        )
        # sql cursor
        self.mycursor = self.connection.cursor()
        
        # combo box auto_fill
        self.ui.site_picker.currentTextChanged.connect(self.autofill_fields_site)
        self.ui.perusahaan_picker.currentTextChanged.connect(self.autofill_fields_perusahaan)
        
        # clicked functions
        self.ui.icon_only_sidebar.hide()
        self.ui.home_btn_full.setChecked(True)
        self.ui.submit.clicked.connect(self.submit_to_db)
        self.ui.home_btn_icon.clicked.connect(lambda: self.load_home_po("", ""))
        self.ui.home_btn_full.clicked.connect(lambda: self.load_home_po("", ""))
        self.ui.refresh_home.clicked.connect(lambda: self.load_home_po("", ""))
        self.ui.update_btn.clicked.connect(self.submit_home_product_page)
        self.ui.update_do_btn.clicked.connect(self.update_do_product_page)
        self.ui.update_do_page_btn.clicked.connect(self.update_do_page)
        self.ui.delete_btn.clicked.connect(self.delete_data)
        self.ui.delete_btn_do_view_more.clicked.connect(self.view_more_do_delete)
        self.ui.delete_btn_do.clicked.connect(self.delete_do_page)
        self.ui.add_po_btn.clicked.connect(self.on_home_btn_icon_toggled)
        self.ui.refresh_btn_2.clicked.connect(self.load_do_table)
        self.ui.search_po.returnPressed.connect(self.search_po)
        self.ui.search_po_btn.clicked.connect(self.search_po)    


        # load data into home page
        self.load_home_po("", "")
        # load data into combo box
        self.load_data_site("")
        self.load_data_perusahaan()

###########################################################################################################################################################
## Search PO function
###########################################################################################################################################################
    def search_po(self):
        search_line = self.ui.search_po.text()
        query = "SELECT purchase_no, purchase_date, product_qty " \
                "FROM purchase_order " \
                "WHERE purchase_no LIKE '%" + search_line + "%'"
        self.load_home_po(query, search_line) 

###########################################################################################################################################################
## Home page functions
###########################################################################################################################################################

    ## for load po data into home page
    def load_home_po(self, query, search):
        self.ui.po_table.clearContents()
        if search == "" or search is None :
            query = "SELECT purchase_no, purchase_date, product_qty FROM purchase_order"
        self.mycursor.execute(query)
        result = self.mycursor.fetchall()
        self.ui.po_table.setRowCount(len(result) - len(result))  # Clear existing rows
        for row in result:
            purchase_no, purchase_date, product_qty = row

            ## view more button
            view_more_btn = QPushButton(self.ui.po_table)
            make_do_btn = QPushButton(self.ui.po_table)
            make_do_btn.setText("Make DO")
            view_more_btn.setText("View More")
            icon = QtGui.QIcon()
            icon.addPixmap(QtGui.QPixmap(":/Green/assets/icon/green/zoom-in.svg"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
            icon.addPixmap(QtGui.QPixmap(":/Green/assets/icon/white/zoom-in.svg"), QtGui.QIcon.Normal, QtGui.QIcon.On)
            view_more_btn.setIcon(icon)

            # Connect button to a lambda that captures row data
            view_more_btn.clicked.connect(lambda _, po=purchase_no, date=purchase_date, qty=product_qty: self.view_more_page(po, date, qty))
            make_do_btn.clicked.connect(lambda _, po=purchase_no: self.make_do(po))


            # main function
            tableIndex = result.index(row)
            self.ui.po_table.insertRow(tableIndex)

            self.ui.po_table.setCellWidget(tableIndex, 0, view_more_btn)
            self.ui.po_table.setCellWidget(tableIndex, 4, make_do_btn)
            self.ui.po_table.setItem(tableIndex, 1, QtWidgets.QTableWidgetItem(str(row[0])))
            self.ui.po_table.setItem(tableIndex, 2, QtWidgets.QTableWidgetItem(str(row[1])))
            self.ui.po_table.setItem(tableIndex, 3, QtWidgets.QTableWidgetItem(str(row[2])))


    def view_more_page(self, po_no, po_date, product_qty):
        self.ui.stackedWidget.setCurrentIndex(1)
        self.ui.po_no_head.setText(str(po_no))
        self.ui.po_date_head.setText(str(po_date))
        self.home_product_page(po_no, product_qty)

        self.ui.add_product_btn.clicked.connect(lambda _, purchaseNo=po_no: self.open_product_window(purchaseNo))

    def home_product_page(self, po_no, product_qty):
        self.ui.products_table.clearContents()
        self.ui.product_qty_head.setText(str(product_qty))
        query = "SELECT ppo.product_id, p.material_no, p.description, ppo.qty, p.unit_of_measurement, ppo.unit_price, ppo.currency , ppo.disc, ppo.total FROM product as p LEFT JOIN product_purchase_order AS ppo ON p.product_id = ppo.product_id WHERE ppo.purchase_no = %s"
        self.mycursor.execute(query, (po_no,))
        result = self.mycursor.fetchall()
        self.ui.products_table.setRowCount(len(result) - len(result))  # Clear existing rows

        self.ui.products_table.hideColumn(0)
        self.ui.products_table.setColumnWidth(1, 25)
        self.ui.products_table.setColumnWidth(5, 50)
        
        for row in result:
            check_btn = QCheckBox(self.ui.products_table)
            check_btn.setObjectName("check_btn")

            print(row)

            # main function
            tableIndex = result.index(row)
            self.ui.products_table.insertRow(tableIndex)
            
            self.ui.products_table.setItem(tableIndex, 0, QtWidgets.QTableWidgetItem(str(row[0])))


            self.ui.products_table.setCellWidget(tableIndex, 1, check_btn)
            self.ui.products_table.setItem(tableIndex, 2, QtWidgets.QTableWidgetItem(str(row[1])))
            self.ui.products_table.setItem(tableIndex, 3, QtWidgets.QTableWidgetItem(str(row[2])))
            self.ui.products_table.setItem(tableIndex, 4, QtWidgets.QTableWidgetItem(str(row[3])))
            self.ui.products_table.setItem(tableIndex, 5, QtWidgets.QTableWidgetItem(str(row[4])))
            self.ui.products_table.setItem(tableIndex, 6, QtWidgets.QTableWidgetItem(str(row[5])))
            self.ui.products_table.setItem(tableIndex, 9, QtWidgets.QTableWidgetItem(str(row[8])))
            if str(row[6]) == None or str(row[6]) == "None":
                self.ui.products_table.setItem(tableIndex, 7, QtWidgets.QTableWidgetItem("IDR"))
            else:
                self.ui.products_table.setItem(tableIndex, 7, QtWidgets.QTableWidgetItem(str(row[6])))
            if str(row[7]) == "0":
                self.ui.products_table.setItem(tableIndex, 8, QtWidgets.QTableWidgetItem("0"))
            else:
                self.ui.products_table.setItem(tableIndex, 8, QtWidgets.QTableWidgetItem(str(row[7])))
        
        delegate = ReadOnlyDelegate(self.ui.products_table)
        self.ui.products_table.setItemDelegateForColumn(2, delegate)
        self.ui.products_table.setItemDelegateForColumn(3, delegate)
        self.ui.products_table.setItemDelegateForColumn(5, delegate)
        self.ui.products_table.setItemDelegateForColumn(9, delegate)
    
    def submit_home_product_page(self):
        po_no = self.ui.po_no_head.text()
        product_qty = self.ui.product_qty_head.text()
        query = "UPDATE `product_purchase_order` SET `qty`= %s,`disc`= %s,`unit_price`= %s, `total`= %s, currency = %s WHERE product_id = %s AND purchase_no = %s"
        for row in range(self.ui.products_table.rowCount()):
            total = 0
            product_id = self.ui.products_table.item(row, 0).text()
            qty = self.ui.products_table.item(row, 4).text()
            unitP = self.ui.products_table.item(row, 6).text()
            currency = self.ui.products_table.item(row, 7).text()
            disc = self.ui.products_table.item(row, 8).text()
            try:
                total = (int(unitP) * (1-int(disc) / 100)) * int(qty)
            except ValueError:
                print("Error Input, Please input all empty cell with number on row", row+1)

            try:
                self.mycursor.execute(query, (qty, disc, unitP, total, currency, product_id, po_no))
                self.connection.commit()
                print("Data Updated!")
            except Exception as e:
                print("Error updating! : ", e)
        
        self.home_product_page(po_no, product_qty)

    def delete_data(self):
        po_no = self.ui.po_no_head.text()
        product = int(self.ui.product_qty_head.text())

        for row in range(self.ui.products_table.rowCount()):
            checked_item = self.ui.products_table.cellWidget(row, 1)
            material_no = self.ui.products_table.item(row, 2).text()
            query = "SELECT product_id FROM product WHERE material_no LIKE '%" + material_no +"%'"
            if isinstance(checked_item, QCheckBox) and checked_item.isChecked():
                self.mycursor.execute(query)
                data = self.mycursor.fetchone() + (po_no,)
                
                self.mycursor.execute("DELETE FROM product_purchase_order WHERE product_id = %s AND purchase_no = %s", data)
                self.connection.commit()

                product = product - 1
        
        self.mycursor.execute("UPDATE purchase_order SET product_qty = %s WHERE purchase_no = %s ", (product, po_no))
        self.connection.commit()

        self.home_product_page(po_no, product)

###########################################################################################################################################################
## for running queries (select, show, etc) || stacked widget change
###########################################################################################################################################################

    ## for running queries (select, show, etc)
    def run_query(self, query, params=None, fetch="one", dictionary=True):
        """
        Executes a query safely using a fresh cursor.
        
        Args:
            query (str): SQL query to execute.
            params (tuple): Parameters for the query.
            fetch (str): "one", "all", or "none".
            dictionary (bool): Whether to return results as dictionaries.

        Returns:
            Query result or None.
        """
        with self.connection.cursor(dictionary=dictionary, buffered=True) as cursor:
            cursor.execute(query, params or ())
            
            if fetch == "one":
                return cursor.fetchone()
            elif fetch == "all":
                return cursor.fetchall()
            else:
                return None

    ## for executing queries (insert, update, delete, etc)
    def execute_query(self, query, params=None):
        with self.connection.cursor(dictionary=True, buffered=True) as cursor:
            cursor.execute(query, params or ())
            self.connection.commit()
            return cursor.lastrowid

    ## Change QPushButton Checkable status when stackedWidget index Changed
    def on_stackedWidget_currentChanged(self, index):
        btn_list = self.ui.icon_only_sidebar.findChildren(QPushButton) \
                    + self.ui.full_sidebar.findChildren(QPushButton)
        
        for btn in btn_list:
            if index in [2, 3]:
                btn.setAutoExclusive(False)
                btn.setChecked(False)
            else:
                btn.setAutoExclusive(True)

    ## function for changing pages
    # home button
    def on_home_btn_full_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(0)
    def on_home_btn_icon_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(0)
    
    # add prodcut button
    '''
    def on_add_product_btn_full_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(4)
    def on_add_product_btn_icon_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(4)
    ''' 
    
    # add Purchase Order button
    def on_add_po_btn_full_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(5)
    def on_add_po_btn_icon_toggled(self):
        self.ui.stackedWidget.setCurrentIndex(5)
    
    # make Delivery Order button
    def on_make_do_btn_full_toggled(self):
        self.load_do_table()
        self.ui.stackedWidget.setCurrentIndex(6)
    def on_make_do_btn_icon_toggled(self):
        self.load_do_table()
        self.ui.stackedWidget.setCurrentIndex(6)

###########################################################################################################################################################
# Purchase Order Form
###########################################################################################################################################################

    # submit button
    def submit_to_db(self):
        nama_site = self.ui.nama_site.text().upper()
        alamat_site = self.ui.alamat_site.toPlainText()
        nama_perusahaan = self.ui.nama_perusahaan.text().upper()
        singkatan_perusahaan = self.ui.singkatan_perusahaan.text().upper()
        npwp = self.ui.npwp_perusahaan.text()
        no_telp_perusahaan = self.ui.telp_perusahaan.text()
        alamat_perusahaan = self.ui.alamat_perusahaan.toPlainText()
        po_no = self.ui.po_no.text()
        po_date = self.ui.po_date.selectedDate().toString("yyyy-MM-dd")

        check_sql = "SELECT pt_id FROM pt_info WHERE pt_name = %s"
        check_val = (nama_perusahaan,)
        result = self.run_query(check_sql, check_val, fetch="one")
        if result:
            pt_id = result["pt_id"]
            print("Perusahaan sudah ada, menggunakan ID:", pt_id)
        else:
            # SQL query to insert data into pt_info
            sql = "INSERT INTO pt_info (pt_name, pt_name_short, pt_alamat, phone, NPWP) VALUES (%s, %s, %s, %s, %s)"
            values = (nama_perusahaan, singkatan_perusahaan, alamat_perusahaan, no_telp_perusahaan, npwp)
            
            try:
                pt_id = self.execute_query(sql, values)
                print("Data inserted successfully!")
            except Exception as e:
                print("Error inserting data pt :", e)
        

        check_sql = "SELECT site_id FROM site_info WHERE site_name = %s AND pt_id = %s"
        check_val = (nama_site, pt_id)
        result = self.run_query(check_sql, check_val, fetch="one")
        if result:
            site_id = result["site_id"]
            print("site sudah ada, menggunakan ID:", site_id)
        else:
            # SQL query to insert data into site_info
            sql = "INSERT INTO site_info (pt_id, site_name, site_alamat) VALUES (%s, %s, %s)"
            values = (pt_id, nama_site, alamat_site)

            try: 
                site_id = self.execute_query(sql, values)
                print("Data inserted successfully!")
            except Exception as e:
                print("Error inserting data site :", e)


        # SQL query to insert data into PO table
        sql = "INSERT INTO purchase_order (purchase_no, purchase_date, pt_id, site_id) VALUES (%s, %s, %s, %s)"
        values = (po_no, po_date, pt_id, site_id)

        try: 
            cursor = self.mycursor
            cursor.execute(sql, values)
            self.connection.commit()
            print("Data inserted successfully!")
        except Exception as e:
            print("Error inserting data po :", e)

        self.load_data_perusahaan()
        self.load_data_site("")
        self.ui.po_no.clear()

    

    # load data into combo box
    def load_data_site(self, pt_id):
        if pt_id != "":
            site_name = self.run_query("SELECT site_name, pt.pt_name_short FROM site_info AS s LEFT JOIN pt_info AS pt on s.pt_id = pt.pt_id WHERE s.pt_id = %s", (pt_id,) ,fetch="all")
        else:
            site_name = self.run_query("SELECT site_name, pt.pt_name_short FROM site_info AS s LEFT JOIN pt_info AS pt on s.pt_id = pt.pt_id", fetch="all")
        self.ui.site_picker.clear()
        self.ui.site_picker.addItem("Select Site")
        for site_name in site_name:
            self.ui.site_picker.addItem(site_name["site_name"] + " : "+ site_name["pt_name_short"])

    def load_data_perusahaan(self):
        perusahaan_name = self.run_query("SELECT pt_name FROM pt_info", fetch="all")
        self.ui.perusahaan_picker.clear()
        self.ui.perusahaan_picker.addItem("Select Perusahaan")
        for pt_name in perusahaan_name:
            self.ui.perusahaan_picker.addItem(pt_name["pt_name"])

    # auto fill function
    def autofill_fields_site(self, site_name):
        site_name = site_name.split(':')[0]

        if site_name == "Select Site":
            self.ui.nama_site.clear()
            self.ui.alamat_site.clear()
            return
        
        query = "SELECT site_name, site_alamat, p.pt_id FROM site_info AS s JOIN pt_info AS p ON s.pt_id = p.pt_id WHERE site_name = %s"
        result = self.run_query(query, (site_name,), fetch="one")
        
        if result is not None:
            self.autofill_fields_perusahaan(result["pt_id"])
        else:
            print("Error : input is not correctly made")

        if result:
            self.ui.nama_site.setText(str(result["site_name"]))
            self.ui.alamat_site.setText(str(result["site_alamat"]))

    def autofill_fields_perusahaan(self, pt_info):
        if pt_info == "Select Perusahaan":
            self.load_data_site("")
            self.clear_perusahaan_fields()
            return
        
        query = "SELECT pt_id, pt_name, pt_name_short, pt_alamat, phone, npwp FROM pt_info WHERE pt_name = %s OR pt_id = %s"
        result = self.run_query(query, (pt_info, pt_info), fetch="one")

        if result:
            self.ui.nama_perusahaan.setText(str(result["pt_name"]))
            self.ui.singkatan_perusahaan.setText(str(result["pt_name_short"]))
            self.ui.npwp_perusahaan.setText(str(result["npwp"]))
            self.ui.telp_perusahaan.setText(str(result["phone"]))
            self.ui.alamat_perusahaan.setText(str(result["pt_alamat"]))
            self.load_data_site(result["pt_id"])
        else:
            self.clear_perusahaan_fields()
    
    def clear_perusahaan_fields(self):
        self.ui.nama_perusahaan.clear()
        self.ui.singkatan_perusahaan.clear()
        self.ui.npwp_perusahaan.clear()
        self.ui.telp_perusahaan.clear()
        self.ui.alamat_perusahaan.clear()
        
###########################################################################################################################################################
## Delivery Order Page
###########################################################################################################################################################

    def make_do(self, po_no):
        # Open Do no, date window
        self.close()
        self.form = DoForm(po_no)
        self.form.show()

        # load data into table
        self.load_do_table()
        self.ui.stackedWidget.setCurrentIndex(6)

    def load_do_table(self):
        self.ui.do_table.clearContents()

        query = "SELECT invoice_no, delivery_id, delivery_date, purchase_no, product_qty, note, invoice_status FROM delivery_order"
        self.mycursor.execute(query)
        result = self.mycursor.fetchall()
        self.ui.do_table.setRowCount(len(result) - len(result))

        for row in result:
            invoice_no, delivery_no, delivery_date, purchase_no, product_qty, note, invoice_status = row

            # button
            check_btn = QCheckBox(self.ui.do_table)

            view_more_btn = QPushButton(self.ui.do_table)
            view_more_btn.setText("View More")
            icon = QtGui.QIcon()
            icon.addPixmap(QtGui.QPixmap(":/Green/assets/icon/green/zoom-in.svg"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
            icon.addPixmap(QtGui.QPixmap(":/Green/assets/icon/white/zoom-in.svg"), QtGui.QIcon.Normal, QtGui.QIcon.On)
            view_more_btn.setIcon(icon)
        
            self.ui.do_table.hideColumn(0) 

            make_invoice_btn = QPushButton(self.ui.do_table)
            make_invoice_btn.setText("Make Invoice")

            # Connect button to a lambda that captures row data
            view_more_btn.clicked.connect(lambda _, do_no=delivery_no, do_date=delivery_date, po_no=purchase_no: self.view_more_do_page(do_no, do_date, po_no))
            make_invoice_btn.clicked.connect(lambda _, do_no=delivery_no: self.make_invoice(do_no))

            # main function
            tableIndex = result.index(row)
            self.ui.do_table.insertRow(tableIndex)
            
            self.ui.do_table.setCellWidget(tableIndex, 1, check_btn)
            self.ui.do_table.setColumnWidth(1, 25)
            self.ui.do_table.setCellWidget(tableIndex, 2, view_more_btn)
            self.ui.do_table.setCellWidget(tableIndex, 9, make_invoice_btn)
            self.ui.do_table.setItem(tableIndex, 0, QtWidgets.QTableWidgetItem(str(row[0])))
            self.ui.do_table.setItem(tableIndex, 3, QtWidgets.QTableWidgetItem(str(row[1])))
            self.ui.do_table.setItem(tableIndex, 4, QtWidgets.QTableWidgetItem(str(row[2])))
            self.ui.do_table.setItem(tableIndex, 5, QtWidgets.QTableWidgetItem(str(row[3])))
            self.ui.do_table.setItem(tableIndex, 6, QtWidgets.QTableWidgetItem(str(row[4])))
            self.ui.do_table.setItem(tableIndex, 7, QtWidgets.QTableWidgetItem(str(row[5])))
            if str(row[6]) == "0":
                self.ui.do_table.setItem(tableIndex, 8, QtWidgets.QTableWidgetItem("Belum Dibuat"))
            else:
                self.ui.do_table.setItem(tableIndex, 8, QtWidgets.QTableWidgetItem("Sudah Dibuat"))

        delegate = ReadOnlyDelegate(self.ui.do_table)
        self.ui.do_table.setItemDelegateForColumn(0, delegate)
        self.ui.do_table.setItemDelegateForColumn(1, delegate)
        self.ui.do_table.setItemDelegateForColumn(2, delegate)
        self.ui.do_table.setItemDelegateForColumn(3, delegate)
        self.ui.do_table.setItemDelegateForColumn(4, delegate)
        self.ui.do_table.setItemDelegateForColumn(5, delegate)
        self.ui.do_table.setItemDelegateForColumn(6, delegate)
        self.ui.do_table.setItemDelegateForColumn(8, delegate)
        self.ui.do_table.setItemDelegateForColumn(9, delegate)

    def update_do_page(self):
        for row in range(self.ui.do_table.rowCount()):
            do_note = self.ui.do_table.item(row, 7).text()
            do_no = self.ui.do_table.item(row, 3).text()

            try:
                self.mycursor.execute("UPDATE delivery_order SET note = %s WHERE delivery_id = %s", (do_note, do_no))
                print("Do Data Updated!")
            except Exception as e:
                print("Error updating do! : ", e)


    def delete_do_page(self):
        # check which product is checked and delete it
        for row in range(self.ui.do_table.rowCount()):
            checked_item = self.ui.do_table.cellWidget(row, 1)
            do_no = self.ui.do_table.item(row, 3).text()
            if isinstance(checked_item, QCheckBox) and checked_item.isChecked():
                
                self.mycursor.execute("DELETE do, ppo " \
                                    "FROM delivery_order as do " \
                                    "JOIN product_purchase_order as ppo ON ppo.delivery_no = do.delivery_id " \
                                    "WHERE delivery_id = %s", (do_no,))
                self.connection.commit()

                self.mycursor.execute("DELETE FROM delivery_order WHERE delivery_id = %s", (do_no,))
                self.connection.commit()

        self.load_do_table()
###########################################################################################################################################################
## Make Invoice
###########################################################################################################################################################

    def make_invoice(self, do_no):
        self.mycursor.execute("SELECT po.purchase_no, site_id, pt_id " \
                            "FROM delivery_order as do " \
                            "JOIN purchase_order as po ON po.purchase_no = do.purchase_no " \
                            "WHERE delivery_id = %s", (do_no,))
        data = self.mycursor.fetchone()
        po_no = data[0]
        site_id = data[1]
        pt_id = data[2]

        self.mycursor.execute("UPDATE delivery_order SET invoice_status = 1 WHERE delivery_id = %s", (do_no,))
        self.connection.commit()

        self.mycursor.execute("SELECT i.invoice_no " \
                            "FROM delivery_order AS do " \
                            "JOIN invoice AS i ON i.delivery_no = do.delivery_id " \
                            "WHERE delivery_id = %s", (do_no,))
        proof = self.mycursor.fetchone()

        try :
            self.mycursor.execute("INSERT INTO invoice (purchase_no, ship_to_site_id, bill_to_pt_id, delivery_no) VALUES (%s, %s, %s, %s)", (po_no, site_id, pt_id, do_no))
            self.connection.commit()
            print("Data invoice inserted successfully!")
        except Exception as e:
            print("Error inserting data invoice :", e)


        self.invoiceForm = InvoiceForm(do_no)
        self.invoiceForm.show()
        self.close()

###########################################################################################################################################################
## Delivery Order View More Page
###########################################################################################################################################################

    def view_more_do_page(self, do_no, do_date, po_no):
        self.ui.stackedWidget.setCurrentIndex(7)
        self.ui.do_view_more_table.clearContents()

        self.ui.do_no_view_more.setText(str(do_no))
        self.ui.po_no_view_more.setText(str(po_no))
        self.ui.do_date_view_more.setText(str(do_date))

        query = "SELECT ppo.product_id, material_no, description, product, qty, p.unit_of_measurement, remark " \
                "FROM product_purchase_order AS ppo " \
                "JOIN delivery_order AS do ON do.delivery_id = ppo.delivery_no " \
                "JOIN product AS p ON p.product_id = ppo.product_id " \
                "WHERE do.delivery_id = %s"
        self.mycursor.execute(query, (do_no,))
        result = self.mycursor.fetchall()

        self.ui.do_view_more_table.setRowCount(len(result) - len(result)) ## fix null table

        for row in result:
            check_btn = QCheckBox(self.ui.do_view_more_table)
            check_btn.setObjectName("check_btn")

            self.ui.do_view_more_table.hideColumn(0)
            self.ui.do_view_more_table.setColumnWidth(1, 25)

            ## Load to table function
            tableIndex = result.index(row)
            self.ui.do_view_more_table.insertRow(tableIndex)

            self.ui.do_view_more_table.setCellWidget(tableIndex, 1, check_btn)
            self.ui.do_view_more_table.setItem(tableIndex, 0, QtWidgets.QTableWidgetItem(str(row[0])))
            self.ui.do_view_more_table.setItem(tableIndex, 2, QtWidgets.QTableWidgetItem(str(row[1])))
            self.ui.do_view_more_table.setItem(tableIndex, 3, QtWidgets.QTableWidgetItem(str(row[2])))
            self.ui.do_view_more_table.setItem(tableIndex, 4, QtWidgets.QTableWidgetItem(str(row[3])))
            self.ui.do_view_more_table.setItem(tableIndex, 5, QtWidgets.QTableWidgetItem(str(row[4])))
            self.ui.do_view_more_table.setItem(tableIndex, 6, QtWidgets.QTableWidgetItem(str(row[5])))
            self.ui.do_view_more_table.setItem(tableIndex, 7, QtWidgets.QTableWidgetItem(str(row[6])))

    def view_more_do_delete(self):
        do_no = self.ui.do_no_view_more.text()
        do_date = self.ui.do_date_view_more.text()
        self.mycursor.execute("SELECT product_qty FROM delivery_order WHERE delivery_id = %s", (do_no,))
        product = int(self.mycursor.fetchone()[0])

        # check which product is checked and delete it
        for row in range(self.ui.do_view_more_table.rowCount()):
            checked_item = self.ui.do_view_more_table.cellWidget(row, 1)
            material_no = self.ui.do_view_more_table.item(row, 2).text()
            query = "SELECT product_id FROM product WHERE material_no LIKE '%" + material_no +"%'"
            if isinstance(checked_item, QCheckBox) and checked_item.isChecked():
                self.mycursor.execute(query)
                data = self.mycursor.fetchone() + (do_no,)
                
                self.mycursor.execute("DELETE FROM product_purchase_order WHERE product_id = %s AND delivery_no = %s", data)
                self.connection.commit()

                product = product - 1
        
        self.mycursor.execute("UPDATE delivery_order SET product_qty = %s WHERE delivery_id = %s ", (product, do_no))
        self.connection.commit()

        self.view_more_do_page(do_no, do_date, product)

    def update_do_product_page(self):
        do_no = self.ui.do_no_view_more.text()
        do_date = self.ui.do_date_view_more.text()
        po_no = self.ui.po_no_view_more.text()
        query = "UPDATE product_purchase_order SET qty = %s, product = %s, remark = %s WHERE product_id = %s AND delivery_no = %s"
        for row in range(self.ui.do_view_more_table.rowCount()):
            product_id = self.ui.do_view_more_table.item(row, 0).text()
            product = self.ui.do_view_more_table.item(row, 4).text()
            qty = self.ui.do_view_more_table.item(row, 5).text()
            remark = self.ui.do_view_more_table.item(row, 7).text()

            try:
                self.mycursor.execute(query, (qty, product, remark, product_id, do_no))
                self.connection.commit()
                print("Data Updated!")
            except Exception as e:
                print("Error updating! : ", e)
        
        auto = auto_input_excel()
        auto.do_auto(do_no, "")
        
        self.view_more_do_page(do_no, do_date, po_no)


###########################################################################################################################################################
## Delivery Order View More Page
###########################################################################################################################################################

    # Open Product Window
    def open_product_window(self, po_no):
        self.product_window = ProductWindow(po_no, "Purchase No :")
        self.product_window.show()

###########################################################################################################################################################
###########################################################################################################################################################

class DoForm(QWidget):
    def __init__(self, po_no):
        super(DoForm, self).__init__()
        #sql connection
        self.connection = mysql.connector.connect (
            host="localhost",
            user="root",
            password="",
            database="database_tdb"
        )
        # sql cursor
        self.mycursor = self.connection.cursor()
        
        self.ui = DoFormUi()
        self.ui.setupUi(self)
        self.setWindowTitle("Delivery Order Form")

        self.ui.warning.hide()
        self.ui.warning.setStyleSheet("color: red;")

        self.ui.do_date_edit.setDate(QtCore.QDate.currentDate())

        self.ui.do_no_lineText.returnPressed.connect(lambda: self.checking(po_no))
        self.ui.submit_btn.clicked.connect(lambda: self.checking(po_no))
    
    def checking(self, po_no):
        if self.ui.do_no_lineText.text() != "":
            self.get_data(po_no)
        else:
            self.ui.warning.show() 

    def get_data(self, po_no):
        do_no = self.ui.do_no_lineText.text()
        do_date = self.ui.do_date_edit.date().toString("yyyy-MM-dd")

        result = self.check_do_no(do_no, po_no)
        if result:
            query = "UPDATE delivery_order SET purchase_no = %s, delivery_date = %s WHERE delivery_id = %s"
        else:
            query = "INSERT INTO delivery_order (purchase_no, delivery_date, delivery_id) VALUE (%s, %s, %s)"

        try:
            self.mycursor.execute(query, (po_no, do_date, do_no))
            self.connection.commit()
        except Exception as e:
            print("Error inserting data do : ", e)
        
        auto_insert = auto_input_excel()
        auto_insert.do_auto(do_no, "bp")

        self.product_window = ProductWindow(do_no, "Delivery No :")
        self.product_window.show()
        '''
        self.mainWindow = MainWindow()
        self.mainWindow.on_make_do_btn_full_toggled()
        self.mainWindow.ui.make_do_btn_full.setChecked(True)
        self.mainWindow.ui.make_do_btn_icon.setChecked(True)
        self.mainWindow.show()
        '''
        self.close()

    def check_do_no(self, do_no, po_no):
        self.mycursor.execute("SELECT delivery_id FROM delivery_order WHERE delivery_id = %s AND purchase_no = %s", (do_no, po_no))
        result = self.mycursor.fetchone()
        if result:
            return True
        else:
            return False


class ProductWindow(QWidget):
    def __init__(self, po_no, label_text):
        super(ProductWindow, self).__init__()

        # sql connection
        self.connection = mysql.connector.connect (
            host="localhost",
            user="root",
            password="",
            database="database_tdb"
        )
        # sql cursor
        self.mycursor = self.connection.cursor()

        # Load the UI from the product_ui module
        self.ui = Ui_Form()
        self.ui.setupUi(self)
        self.setWindowTitle("Product Form")


        self.ui.no_label.setText(label_text)
        self.ui.po_no_label.setText(str(po_no))
        
        # load data from database
        self.load_data("")

        self.ui.products_table.setHorizontalHeaderLabels(["", "Material No", "Nama Produk", "UoM"])


        # push button connect
        self.ui.search_btn.clicked.connect(self.search_product)
        self.ui.submit_btn.clicked.connect(lambda _, purchaseNo=po_no: self.submit_product(purchaseNo))
        self.ui.add_row_btn.clicked.connect(self.add_row)
        self.ui.search_bar.returnPressed.connect(self.search_product)

    def search_product(self):
        search_text = self.ui.search_bar.text()
        if self.ui.no_label.text() == "Purchase No :":
            query = "SELECT material_no, description, unit_of_measurement " \
                    "FROM product " \
                    "WHERE material_no LIKE '%" + search_text + "%' OR description LIKE '%" + search_text + "%'"
        else:
            query = "SELECT p.material_no, p.description, p.unit_of_measurement " \
                    "FROM product AS p " \
                    "JOIN product_purchase_order AS ppo ON p.product_id = ppo.product_id " \
                    "JOIN delivery_order AS do ON ppo.purchase_no = do.purchase_no " \
                    "WHERE (p.material_no LIKE '%" + search_text + "%' " \
                    "OR p.description LIKE '%" + search_text + "%') AND do.delivery_id LIKE %s"
        self.ui.search_bar.clear()
        self.load_data(query)

    def load_data(self, query2):
        # retrieve data from database
        no = self.ui.po_no_label.text()
        label = self.ui.no_label.text()
        if label == "Purchase No :":
            query = "SELECT material_no, description, unit_of_measurement " \
                    "FROM product"
            if query2 == "":
                self.mycursor.execute(query)
            else:
                self.mycursor.execute(query2)
        else:
            query = "SELECT p.material_no, p.description, p.unit_of_measurement " \
                    "FROM product p " \
                    "JOIN product_purchase_order ppo on p.product_id = ppo.product_id " \
                    "JOIN delivery_order AS do ON ppo.purchase_no = do.purchase_no " \
                    "WHERE do.delivery_id LIKE %s"
            if query2 == "":
                self.mycursor.execute(query, (no,))
            else:
                self.mycursor.execute(query2, (no,))
        
        result = self.mycursor.fetchall()
        self.ui.products_table.setRowCount(len(result) - len(result))  # Clear existing rows

        for row in result:
            # check mark button
            check_btn = QCheckBox(self.ui.products_table)
            check_btn.setObjectName("check_btn")

            # main function
            tableIndex = result.index(row)
            self.ui.products_table.insertRow(tableIndex)

            self.ui.products_table.setColumnWidth(0, 25)
            self.ui.products_table.setColumnWidth(3, 50)
            self.ui.products_table.setCellWidget(tableIndex, 0, check_btn)
            self.ui.products_table.setItem(tableIndex, 1, QtWidgets.QTableWidgetItem(str(row[0])))
            self.ui.products_table.setItem(tableIndex, 2, QtWidgets.QTableWidgetItem(str(row[1])))
            self.ui.products_table.setItem(tableIndex, 3, QtWidgets.QTableWidgetItem(str(row[2])))
            tableIndex += 1
    
    def add_row(self):
        check_btn = QCheckBox(self.ui.products_table)
        check_btn.setObjectName("check_btn")

        current_row = self.ui.products_table.rowCount()
        self.ui.products_table.insertRow(current_row)

        self.ui.products_table.setCellWidget(current_row, 0, check_btn)
        check_btn.setChecked(True)

    def submit_product(self, po_no):
        product = 0

        for row in range(self.ui.products_table.rowCount()):
            check_btn = self.ui.products_table.cellWidget(row, 0)

            if isinstance(check_btn, QCheckBox) and check_btn.isChecked():
                material_no = (self.ui.products_table.item(row, 1).text())
                description = (self.ui.products_table.item(row, 2).text())
                uom = (self.ui.products_table.item(row, 3).text())
            else:
                continue
        
            check_query = "SELECT product_id FROM product WHERE material_no = %s AND description = %s AND unit_of_measurement = %s"
            check_val = material_no, description, uom
            self.mycursor.execute(check_query, check_val)
            result = self.mycursor.fetchone()
            ## for po
            if self.ui.no_label.text() == "Purchase No :":
                if result:
                    product_id = result[0]
                    print("Product sudah ada, dengan ID Product :", product_id)
                    try:
                        self.mycursor.execute("SELECT product_id FROM product_purchase_order WHERE purchase_no = %s AND product_id = %s", (po_no, product_id))
                        result2 = self.mycursor.fetchone()
                        if result2:
                            print("product already added!")
                        else:
                            self.mycursor.execute("INSERT INTO product_purchase_order (purchase_no, product_id) VALUES (%s, %s)", (po_no, product_id))
                            print("po_no updated!")
                            self.connection.commit()
                            product += 1
                    except Exception as e:
                        print("Error inserting data product po :", e)

                else:    
                    try: 
                        self.mycursor.execute("INSERT INTO product (material_no, description, unit_of_measurement) VALUES (%s, %s, %s)", (material_no, description, uom))
                        self.connection.commit()
                        product_id = self.mycursor.lastrowid
                        print("Data inserted successfully with ID:", product_id)
                        self.mycursor.execute("INSERT INTO product_purchase_order (purchase_no, product_id) VALUES (%s, %s)", (po_no, product_id))
                        self.connection.commit()
                        print("po_no updated!")
                        product += 1
                    except Exception as e:
                        print("Error inserting data product :", e)

            ## for do        
            else :
                self.mycursor.execute("SELECT purchase_no" \
                                    " FROM delivery_order " \
                                    "WHERE delivery_id = %s", (po_no,))
                po_no_result = self.mycursor.fetchone()[0]
                if result:
                    product_id = result[0]
                    print("Product sudah ada, dengan ID Product :", product_id)
                else:    
                    try: 
                        self.mycursor.execute("INSERT INTO product (material_no, description, unit_of_measurement) VALUES (%s, %s, %s)", (material_no, description, uom))
                        self.connection.commit()
                        product_id = self.mycursor.lastrowid
                        print("Data inserted successfully with ID:", product_id)
                    except Exception as e:
                        print("Error inserting data product :", e)

                self.mycursor.execute("SELECT qty " \
                                    "FROM product_purchase_order " \
                                    "WHERE product_id = %s AND purchase_no = %s", (product_id, po_no_result))
                try:
                    qty_result = self.mycursor.fetchone()[0]
                except Exception as e:
                    print("Error fetching qty:", e)
                    qty_result = 0
                
                try:
                    self.mycursor.execute("SELECT product_id " \
                                        "FROM product_purchase_order " \
                                        "WHERE delivery_no = %s AND product_id = %s", (po_no, product_id))
                    result2 = self.mycursor.fetchone()
                    if result2:
                        print("product already added!")
                    else:
                        self.mycursor.execute("INSERT INTO product_purchase_order (delivery_no, product_id, qty, purchase_no) VALUES (%s, %s, %s, %s)", (po_no, product_id, qty_result, po_no_result))
                        self.connection.commit()
                        print("do_no updated!")
                        product += 1
                except Exception as e:
                    print("Error inserting data product do :", e)
                

    
        self.product_qty_plus(po_no, product)
        mainWindow = MainWindow()
        if self.ui.no_label.text() == "Delivery No :":
            mainWindow.on_make_do_btn_full_toggled()
            mainWindow.ui.make_do_btn_full.setChecked(True)
            mainWindow.ui.make_do_btn_icon.setChecked(True)
        mainWindow.show()
    
    def product_qty_plus(self, po_no, qty):
        if self.ui.no_label.text() == "Purchase No :":
            query = "UPDATE purchase_order SET product_qty = product_qty + %s WHERE purchase_no = %s"
        elif self.ui.no_label.text() == "Delivery No :":
            query = "UPDATE delivery_order SET product_qty = product_qty + %s WHERE delivery_id = %s"
        
        try:
            self.mycursor.execute(query, (qty, po_no))
            self.connection.commit()
        except Exception as e:
            print("Error updating data:", e)
        
class InvoiceForm(QWidget):
    def __init__(self, do_no):
        super(InvoiceForm, self).__init__()
        #sql connection
        self.connection = mysql.connector.connect (
            host="localhost",
            user="root",
            password="",
            database="database_tdb"
        )
        # sql cursor
        self.mycursor = self.connection.cursor()
        
        self.ui = InvoiceFormUi()
        self.ui.setupUi(self)
        self.setWindowTitle("Invoice Form")

        ## defaulting widget
        self.ui.invoice_date.setDate(QtCore.QDate.currentDate())
        self.ui.top_date.setDate(QtCore.QDate.currentDate())
        self.ui.ship_date.setDate(QtCore.QDate.currentDate())
        self.ui.pbbkb.setText("-")
        self.ui.pph.setText("-")
        self.ui.ship_via.setText("-")
        self.ui.invoice_desc.setText("-")

        ##adding the combo box item
        self.ui.ship_date_combo_box.addItems(["-", "Pilih Tanggal"])


        ## hide ship date edit
        self.ui.ship_date.hide()

        ## combo picker
        self.ui.ship_date_combo_box.currentTextChanged.connect(self.combo_box_changed)

        ## Push Button connect
        self.ui.submit_btn.clicked.connect(lambda: self.submit_invoice(do_no))
        self.ui.refresh_btn.clicked.connect(self.refresh_func)

    def refresh_func(self):
        self.ui.ship_date.hide()
        self.ui.ship_date_combo_box.show()
        self.ui.ship_date_combo_box.setCurrentIndex(0)

    def combo_box_changed(self, text):
        if text == "-":
            pass
        else :
            self.ui.ship_date_combo_box.hide()
            self.ui.ship_date.show()

    def submit_invoice(self, do_no):
        invoice_no = self.ui.invoice_no_edit.text()
        invoice_date = self.ui.invoice_date.date().toString("yyyy-MM-dd")
        terms_of_payment = self.ui.top_date.date().toString("yyyy-MM-dd")
        ship_via = self.ui.ship_via.text()
        description = self.ui.invoice_desc.toPlainText()

        ## pbbkb and pph check
        if self.ui.pbbkb.text() == "-":
            pbbkb = 0
        else:
            pbbkb = self.ui.pbbkb.text()
        
        if self.ui.pph.text() == "-":
            pph = 0
        else :     
            pph = self.ui.pph.text()       

        ## ship date check
        if self.ui.ship_date_combo_box.currentText() == "-":
            ship_date = "-"
        else:
            ship_date = self.ui.ship_date.date().toString("yyyy-MM-dd")

        try:
            self.mycursor.execute("UPDATE invoice " \
                                "SET invoice_no = %s, invoice_date = %s, term_of_payment = %s, " \
                                "shipping_via = %s, description = %s, ship_date = %s " \
                                "WHERE delivery_no = %s", (invoice_no, invoice_date, terms_of_payment, ship_via, description, ship_date, do_no))
            self.connection.commit()
            self.mycursor.execute("UPDATE delivery_order SET invoice_no = %s WHERE delivery_id = %s", (invoice_no, do_no))
            self.connection.commit()
            print("Data invoice updated successfully!")
            self.auto_bridge(do_no)
        except Exception as e:
            print("Error updating data invoice :", e)

        ## close form
        mainWindow = MainWindow()
        mainWindow.show()
        self.close()
    
    def auto_bridge(self, do_no):
        invoice_auto = auto_input_excel()
        invoice_auto.inovice_auto(do_no)

class auto_input_excel():
    def __init__(self):    
        # sql connection
        self.connection = mysql.connector.connect (
            host="localhost",
            user="root",
            password="",
            database="database_tdb"
        )
        # sql cursor
        self.mycursor = self.connection.cursor()


    def do_auto(self, do_no, atasNama):
        ## variable
        rowExcel = 22
        no = 1

        ## load excel file
        do_wb = load_workbook('FORMAT_DO.xlsx')
        do_ws = do_wb.active

        ## getting the data

        # data tujuan
        self.mycursor.execute("SELECT po.purchase_no, pt.pt_name, site.site_name, do.note, do.delivery_date " \
                            "FROM delivery_order AS do " \
                            "JOIN purchase_order AS po ON po.purchase_no = do.purchase_no " \
                            "JOIN site_info AS site ON site.site_id = po.site_id " \
                            "JOIN pt_info AS pt ON pt.pt_id = po.pt_id " \
                            "WHERE do.delivery_id = %s", (do_no,))

        result = self.mycursor.fetchone()

        # product data
        self.mycursor.execute("SELECT p.material_no, p.description, ppo.product, ppo.qty, p.unit_of_measurement, ppo.remark " \
                            "FROM product_purchase_order AS ppo " \
                            "JOIN product AS p on p.product_id = ppo.product_id " \
                            "WHERE delivery_no = %s", (do_no,))
        result2 = self.mycursor.fetchall()

        # insert data product
        for row in result2:
            currentRow = rowExcel + no
            number = "A" + str(currentRow)
            partNo = "C" + str(currentRow)
            desc = "D" + str(currentRow)
            product = "E" + str(currentRow)
            qty = "F" + str(currentRow)
            oum = "G" + str(currentRow)
            remark = "H" + str(currentRow)

            do_ws[number].value = str(no)    
            do_ws[partNo].value = str(row[0])    
            do_ws[desc].value = str(row[1])    
            do_ws[product].value = str(row[2])    
            do_ws[qty].value = str(row[3])    
            do_ws[oum].value = str(row[4])    
            do_ws[remark].value = str(row[5])    

            no += 1


        # insert data tujuan
        do_ws['C10'].value = result[4]
        do_ws['C11'].value = do_no
        do_ws['C12'].value = result[0]
        do_ws['C15'].value = str(result[1])
        do_ws['C16'].value = "Site " + str(result[2])
        do_ws['C17'].value = atasNama
        do_ws['C37'].value = str(result[3])

        do_new_file_name = "DO new.xlsx"

        do_wb.save(do_new_file_name)

    def inovice_auto(self, do_no):
        ## variable
        rowExcel = 17
        no = 1

        ## load excel file
        invoice_wb = load_workbook('FORMAT_INVOICE.xlsx')
        invoice_ws = invoice_wb.active

        ### auto insert function

        ## getting the data
        # data tujuan
        self.mycursor.execute("SELECT i.invoice_date, i.invoice_no, i.shipping_via, i.ship_date, i.purchase_no, i.description, i.bill_to_pt_id, i.ship_to_site_id, i.term_of_payment " \
                            "FROM invoice AS i " \
                            "JOIN delivery_order AS do ON do.invoice_no = i.invoice_no " \
                            "WHERE i.delivery_no = %s", (do_no,))
        result = self.mycursor.fetchone()

        self.mycursor.execute("SELECT site_name, site_alamat " \
                            "FROM site_info " \
                            "WHERE site_id = %s", (str(result[7]),))
        site = self.mycursor.fetchone()

        self.mycursor.execute("SELECT pt_name, pt_name_short, pt_alamat, NPWP " \
                            "FROM pt_info " \
                            "WHERE pt_id = %s",(str(result[6]),))
        pt = self.mycursor.fetchone()

        # data product
        self.mycursor.execute("SELECT p.material_no, p.description, ppo.qty, p.unit_of_measurement, ppo.currency, ppo.unit_price, ppo.disc " \
                            "FROM product_purchase_order AS ppo " \
                            "JOIN product AS p ON p.product_id = ppo.product_id " \
                            "WHERE ppo.purchase_no = %s AND ppo.delivery_no IS NULL", (str(result[4]),))
        result2 = self.mycursor.fetchall()

        ### inserting the data
        # insert data product
        for row in result2:
            currentRow = rowExcel + no
            number = "A" + str(currentRow)
            item = "B" + str(currentRow)
            desc = "F" + str(currentRow)
            qty = "K" + str(currentRow)
            uom = "L" + str(currentRow)
            currency = "N" + str(currentRow)
            unit_price = "P" + str(currentRow)

            invoice_ws[number].value = str(no)
            invoice_ws[item].value = str(row[0])
            invoice_ws[desc].value = str(row[1])
            invoice_ws[qty].value = row[2]
            invoice_ws[uom].value = str(row[3])
            invoice_ws[currency].value = str(row[4])
            invoice_ws[unit_price].value = row[5]
            invoice_ws['S17'].value = str(row[6])

            no += 1

        # insert data tujuan
        invoice_ws['C12'].value = str(pt[0])
        invoice_ws['C13'].value = str(pt[2])
        invoice_ws['C15'].value = str(pt[3])
        invoice_ws['M9'].value = str(pt[0]) + ";JOBSITE:" + str(site[0])
        invoice_ws['M10'].value = str(site[1])
        invoice_ws['M6'].value = "SITE " + str(pt[1]) + " " +str(site[0])

        invoice_ws['K13'].value = str(result[0])
        invoice_ws['N14'].value = str(result[1])
        invoice_ws['R13'].value = str(result[2])
        if str(result[3]) == "None" :
            invoice_ws['T13'].value = "-"
        else :
            invoice_ws['T13'].value = str(result[3])
        invoice_ws['R14'].value = str(result[4])
        invoice_ws['C33'].value = str(result[5])
        invoice_ws['O13'].value = str(result[8])

        invoice_new_file_name = "invoice_new.xlsx"

        invoice_wb.save(invoice_new_file_name)



# execute application
if __name__ == "__main__":
    app = QApplication(sys.argv)

    # loading style file
    with open("style.qss", "r") as style_file:
        style_str = style_file.read()
    app.setStyleSheet(style_str)

    window = MainWindow()
    window.show()

    sys.exit(app.exec())